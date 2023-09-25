from django.shortcuts import render, redirect, HttpResponse
from .models import *
import base64
from django.http import JsonResponse
# import timezone
from django.utils import timezone

def home(request):
    products = Products.objects.all()
    print(products)
    banner = PromoItems.objects.all().last()
    
    if banner:
        banner = banner
        banner = base64.b64encode(banner.promo_image).decode('utf-8')
    # top 5 best offers
    best_offers = products[:5]
    # create a session for the user and give them a unique key
    unique_key = generate_unique_key()
    request.session['unique_key'] = unique_key
    if products:
        print('hi')
    context = {
        'products': products,
        'banner': banner,
        'best_offers': best_offers
    }
    
    return render(request, 'test.html', context)
    # return render(request, 'test.html')
def upload_product(request):
    if request.method == 'POST':
        item_name = request.POST['item_name']
        description = request.POST['description']
        price = request.POST['price']
        stock = request.POST['stock']
        # discount = request.POST.get('discount', False)
        discount = True 
        discount_percentage = request.POST['discount_percentage']
        image = request.FILES['image'].read() if 'image' in request.FILES else None
        category_id = request.POST['category']
        category = ProductCategory.objects.get(pk=category_id)
        
        Products.objects.create(
            item_name=item_name,
            description=description,
            price=price,
            stock=stock,
            discount=discount,
            discount_percentage=discount_percentage,
            image=image,
            category=category
        )
        return redirect('dashboard')

    # categories = ProductCategory.objects.all()
    # return render(request, 'upload_product.html', {'categories': categories})

def upload_promoitems(request):
    # upload to promoitems model
    if request.method == 'POST':
        promo_name = request.POST['promo_name']
        promo_image = request.FILES['promo_image'].read() if 'promo_image' in request.FILES else None
        promo_description = request.POST['promo_description']
        PromoItems.objects.create(
            promo_name=promo_name,
            promo_image=promo_image,
            promo_description=promo_description
        )
        return HttpResponse('Promo Item Uploaded')
    return render(request, 'upload_promo.html')


def product_category(request, product_category):
    if request.method == 'GET':
        # category_name = request.GET['category']
        category_name = product_category
        # in product category, find the id that matches the category name
        category_id = ProductCategory.objects.get(name=category_name)

        products = Products.objects.filter(category=category_id)
        print(products)
        return render(request, 'products_by_category.html', {'products': products, 'category_name': category_name})
    
def product(request, product_id):
    if request.method == 'GET':
        product = Products.objects.get(pk=product_id)
        print(product)
        return render(request, 'product.html', {'product': product})
def generate_unique_key():
    # generate a unique key
    import uuid
    return str(uuid.uuid4())
from django.shortcuts import get_object_or_404
# def add_to_cart(request, product_id):
#     if request.method == 'GET':
#         cart_key = request.session.get('cart_key', None)
#         if not cart_key:
#             cart_key = generate_unique_key()
#             request.session['cart_key'] = cart_key

#         product = get_object_or_404(Products, pk=product_id)
#         existing_cart = Cart.objects.filter(cart_key=cart_key, ordered=False).first()
#         if existing_cart:
#             # Update the existing cart item's quantity instead of creating a new one
#             existing_cart.quantity += 1
#             existing_cart.save()
#         else:
#             Cart.objects.create(cart_key=cart_key, buyer='example', ordered=False, quantity=1, product=product)
#         return HttpResponse('Cart created')
# def add_to_cart(request, product_id):
#     product = Products.objects.get(pk=product_id)
#     cart, created = Cart.objects.get_or_create(user=request.user)
#     cart_item, item_created = CartItem.objects.get_or_create(cart=cart, product=product)
    
#     if not item_created:
#         cart_item.quantity += 1
#         cart_item.save()
    
#     return HttpResponse('Cart created')
from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def add_to_cart(request):
    if request.method == 'POST':
        product_id = request.POST['id']
        quantity = int(request.POST.get('quantity'))  # Default to 1 if quantity is not provided
        print('product id: ', product_id, 'quantity: ', quantity)
        product = Products.objects.get(pk=product_id)
        
        # if user is anonymous
        if not request.user.is_authenticated:
            # Check if the user has a cart in the session or create one
            if 'cart_id' not in request.session:
                cart = AnonymousCart.objects.create()
                request.session['cart_id'] = cart.pk
            else:
                cart_id = request.session['cart_id']
                cart = get_object_or_404(AnonymousCart, pk=cart_id)

            # Add the product to the cart
            cart_item, created = AnonymousCartItem.objects.get_or_create(cart=cart, product=product)
            print('cart item: ', cart_item)
            print('created: ', created)
            if not created:
                cart_item.quantity += quantity
                cart_item.save()
                
            return JsonResponse({'success': True})
        # if user is authenticated
        else:
            cart, created = Cart.objects.get_or_create(user=request.user)
            cart_item, item_created = CartItem.objects.get_or_create(cart=cart, product=product, quantity=quantity)
            
            if not item_created:
                cart_item.quantity += quantity
                cart_item.save()
                
            return JsonResponse({'success': True})
    return JsonResponse({'success': False})

# quantity in cart
def quantity_in_cart(request):
    if request.method == 'GET':
        if request.user.is_authenticated:
            cart = Cart.objects.get(user=request.user)
            cart = cart.cartitem_set.count()
            return JsonResponse({'quantity': cart})
        else:
            print('anonymous user')
            cart_id = request.session.get('cart_id')
            print(cart_id)
            if cart_id:
                cart = AnonymousCart.objects.get(id=cart_id)
                cart = cart.anonymouscartitem_set.count()
                return JsonResponse({'quantity': cart})
            else:
                return JsonResponse({'quantity': 0})
    return JsonResponse({'quantity': 0})

def view_cart(request):
    context = {}
    if request.user.is_authenticated:
        try: 
            cart = Cart.objects.get(user=request.user)
            cart_items = cart.cartitem_set.all()
        except:
            cart_items = []
    else:
        cart_id = request.session.get('cart_id')
        if cart_id:
            cart = AnonymousCart.objects.get(id=cart_id)
            cart_items = cart.anonymouscartitem_set.all()
        else:
            cart_items = []
    try:
        total_cost = cart.total
        context['cart_items'] = cart_items
        context['total_cost'] = total_cost
    except:
        print('No cart')
    # total_cost = cart.total if cart else 0
    
    return render(request, 'view_cart.html', context)
@csrf_exempt
def update_cart(request):
    if request.method == 'POST':
        cart_item_id = request.POST['id']
        quantity = int(request.POST['quantity'])
        if request.user.is_authenticated:
            cart_item = CartItem.objects.get(pk=cart_item_id)
            cart_item.quantity = quantity
            cart_item.save()
            return JsonResponse({'success': True, 'sub_total': cart_item.sub_total, 'total_cost': cart_item.cart.total})
        else:
            cart_item = AnonymousCartItem.objects.get(pk=cart_item_id)
            print(cart_item)
            cart_item.quantity = quantity
            cart_item.save()
        return JsonResponse({'success': True, 'sub_total': cart_item.sub_total, 'total_cost': cart_item.cart.total})
    return JsonResponse({'success': False})

@csrf_exempt
def remove_from_cart(request):
    if request.method == 'POST':
        cart_item_id = request.POST['id']
        if request.user.is_authenticated:
            cart_item = CartItem.objects.get(pk=cart_item_id)
            # update the cart total
            cart = cart_item.cart
            cart.total -= cart_item.sub_total
            cart_item.delete()
            return JsonResponse({'success': True, 'total_cost': cart.total})
        else:
            cart_item = AnonymousCartItem.objects.get(pk=cart_item_id)
            cart_item.delete()
            return JsonResponse({'success': True, 'total_cost': cart.total})
    return JsonResponse({'success': False})


from .models import Order, OrderItem, BuyerDeliveryDetails
def checkout(request):
    if request.method == 'POST':
        
        # get the data from the form to populate the order
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        email = request.POST['email']
        address = request.POST['address']
        phone_number = request.POST['phone_number']
        postal_code = request.POST['postal_code']
        city = request.POST['city']
        country = request.POST['country']
        payment_method = request.POST['payment_method']
        # add the data to the buyer delivery details
        buyer = BuyerDeliveryDetails.objects.create(
            first_name=first_name,
            last_name=last_name,
            email=email,
            address=address,
            phone_number=phone_number,
            postal_code=postal_code,
            city=city,
            country=country,
            payment_method=payment_method
        )
        print(buyer)
        # create an order and populate it with the cart items and the buyer delivery details
        if request.user.is_authenticated:
            cart = Cart.objects.get(user=request.user)
            cart_items = cart.cartitem_set.all()
            order = Order.objects.create(buyer_id_order=request.user, buyer=buyer)

            for item in cart_items:
                OrderItem.objects.create(order=order, product=item.product, quantity=item.quantity)
            # delete the cart
            for item in cart_items:
                item.product.sold += item.quantity
                item.product.save()
            cart.delete()
            # increase the sold count of the product
            
            return redirect('view_orders')
        else:
            cart = AnonymousCart.objects.get(id=request.session['cart_id'])
            cart_items = cart.anonymouscartitem_set.all()
            order = Order.objects.create(buyer=buyer)
            for item in cart_items:
                OrderItem.objects.create(order=order, product=item.product, quantity=item.quantity)
            print(first_name, last_name, email, address, phone_number, postal_code, city)
            # set order session
            for item in cart_items:
                item.product.sold += item.quantity
                item.product.save()
            cart.delete()
            
            request.session['order_id'] = order.id
        # create an order
        return redirect('view_orders')
from django.db.models import Sum
def view_orders(request):
    if request.user.is_authenticated:
        # print(request.user.username)
        orders = Order.objects.filter(buyer_id_order=request.user)
        order_items = OrderItem.objects.filter(order=orders.last())
        # sum the sub total of the order items
        total_cost = order_items.aggregate(Sum('sub_total'))
        # clean the total cost
        total_cost = total_cost['sub_total__sum']
    else:
        print(request.session['order_id'])
        print('anonymous user')
        # return orders for anonymous user
        orders = Order.objects.filter(id= request.session['order_id'])
        order_items = OrderItem.objects.filter(order=orders.last())
        total_cost = order_items.aggregate(Sum('sub_total'))
        # clean the total cost
        total_cost = total_cost['sub_total__sum']
    return render(request, 'view_orders.html', {'orders': orders, 'order_items': order_items, 'total_cost': total_cost})
    
# dashboard view that shows products, carts, orders, sales, etc
from django.contrib.sessions.models import Session
from django.contrib.auth.decorators import login_required
@login_required
def dashboard(request):
    # get all the products
    products = Products.objects.all()
    orders = Order.objects.all()
    # print(orders)
    order_items = OrderItem.objects.all()
    product_counts = products.count()
    active_sessions = Session.objects.filter(expire_date__gte=timezone.now())
    active_session_count = active_sessions.count()
    open_orders = orders.filter(ordered=False).count()

    # get all exceptions in the last 24 hours
    exceptions = ExceptionLog.objects.filter(exception_date__gte=timezone.now() - timezone.timedelta(hours=24))
    exception_count = exceptions.count()
    print(exception_count)

    print(order_items)
    context = {
        'products': products,
        'order_items': orders,
        'open_orders': open_orders,
        'product_counts': product_counts,
        'active_users': active_session_count,
        'exception_count': exception_count,
    }

    return render(request, 'shop/dashboard.html', context)

import json
# search inventory
def search_inventory(request):
    if request.method == 'GET':
        search_term = request.GET['query']
        print(search_term)
        # search for the product
        products = Products.objects.filter(item_name__icontains=search_term)
        # convert the products to json including the image
        products = json.dumps([{
            'id': product.id, 'item_name': product.item_name, 
            'image': base64.b64encode(product.image).decode('utf-8'),
            'price': str(product.price), 
            # 'discount': product.discount,
            # 'discount_percentage': product.discount_percentage,
            'stock': str(product.stock),
            }
            for product in products])
        # print(products)
        return JsonResponse({'products': products})
    
def products_admin(request):
    products = Products.objects.all()
    return render(request, 'shop/products.html', {'products': products})

def add_products_page(request):
    # categories = ProductCategory.objects.all()
    return render(request, 'shop/add_products.html')


def add_category(request):
    pass
    
from django.core.files.base import ContentFile
from openpyxl import load_workbook

def read_image_as_binary(image_path):
    with open(image_path, 'rb') as image_file:
        binary_data = image_file.read()
    return binary_data


def upload_products(request):
    if request.method == 'POST' and request.FILES.get('excel'):
        excel_file = request.FILES['excel']
        wb = load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # name, image_path = row
            # product = Product(
            #     product_name=name,
            # )
            # my model is called Product
            # upload to the model
            product = Products(
                item_name=row[0],
                description=row[1],
                price=row[2],
                stock=row[3],
                discount=row[4],
                discount_percentage=row[5],
                # category=row[6]
            )
            category_name = row[6]
            product_category, created = ProductCategory.objects.get_or_create(name=category_name)
            product.category = product_category

            image_path = row[7]
            binary_image = read_image_as_binary(image_path)
            print(image_path)
            if binary_image:
                product.image = binary_image
                product.save()
            else:
                print(f"Failed to read image for product {product.item_name}")
        return redirect('dashboard')
    return HttpResponse('No file uploaded')

def edit_product(request, pk):
    if request.method == 'GET':
        product = Products.objects.get(id=pk)
        return render(request, 'shop/edit_product.html', {'product': product})
    if request.method == 'POST':
        product = Products.objects.get(id=pk)
        product.item_name = request.POST['name']
        product.description = request.POST['description']
        product.price = request.POST['price']
        product.stock = request.POST['stock']
        product.discount = True 
        product.discount_percentage = request.POST['discount_percentage']
        image = request.FILES.get('image')
        print(image)
        if image:
            product.image = image.read()

        # product category is obtained from options
        category_id = request.POST['category']
        print(category_id)
        product.save()
        return redirect('products')
    
def delete_product(request, pk):
    product = Products.objects.get(id=pk)
    product.delete()
    # return redirect('products')
    return JsonResponse({'success': True})


def admin_orders(request):
    orders = Order.objects.all()
    # buyer of the order
    return render(request, 'shop/orders.html', {'orders': orders})

def admin_order_details(request, pk):
    order = Order.objects.get(id=pk)
    order_items = OrderItem.objects.filter(order=order)

    return render(request, 'shop/order_details.html', {'order_items': order_items})

from itertools import chain
def admin_carts(request):
    # combine the two carts
    carts = []
    # get the anonymous carts
    anonymous_carts = AnonymousCart.objects.all()
    # get the carts
    ver_carts = Cart.objects.all()
    # combine the two carts
    carts = list(chain(anonymous_carts, ver_carts))


    return render(request, 'shop/carts.html', {'carts': carts})

def view_admin_cart(request, pk):
    cart = Cart.objects.get(id=pk)
    cart_items = cart.cartitem_set.all()
    return render(request, 'shop/cart_desc.html', {'cart_items': cart_items})

# when admin clicks deliver, the order status changes to delivered and store it in sales and reduce the stock
def deliver_order(request, pk):
    order = Order.objects.get(id=pk)
    order.ordered = True
    order.save()
    # reduce the stock
    order_items = OrderItem.objects.filter(order=order)
    for item in order_items:
        product = item.product
        product.stock -= item.quantity
        product.save()
    # store the order in sales
    Sale.objects.create(order=order)
    return redirect('orders')

def sales(request):
    sales = Sale.objects.all()
    for sale in sales:
        print(sale.order.buyer_id_order)

    # get the total cost of each sales from the order
    sales_value = []
    for sale in sales:
        order_items = OrderItem.objects.filter(order=sale.order)
        total_cost = order_items.aggregate(Sum('sub_total'))
        # clean the total cost
        total_cost = total_cost['sub_total__sum']
        sales_value.append(total_cost)

    
    # combine sale with sales_value
    # sales = zip(sales, sales_value)

    return render(request, 'shop/sales.html', {'sales': sales, 'sales_items': sales_value})

def product_image(request, pk):
    product = Product.objects.get(id=pk)
    image_data = product.image_binary
    return HttpResponse(image_data, content_type='image/jpeg') 
